"""
Admin API routes for managing session data.

Provides endpoints for:
- Searching sessions by username
- Sorting by date, training type
- Downloading individual files
- Bulk download as ZIP
"""

import json
import zipfile
import io
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Any, Optional
from flask import Blueprint, jsonify, request, send_file, abort

import state
from routes.auth import admin_required
from hrv import compute_lf_power, compute_rmssd

admin_api_bp = Blueprint("admin_api", __name__)


def load_session_data(filepath: Path) -> Optional[Dict[str, Any]]:
    """Load and parse a session JSON file."""
    try:
        with open(filepath, "r", encoding="utf-8") as f:
            data = json.load(f)
        return data
    except Exception:
        return None


def detect_training_type(data: Dict[str, Any]) -> str:
    """Detect training type from session data."""
    # Check for Buteyko-specific fields
    if "bolt_seconds" in data or "round_results" in data or "buteyko_config" in data:
        return "buteyko"
    # Check for PRBF/adaptive experiment
    if "prbf_test" in data or data.get("config", {}).get("prbf_mode"):
        return "prbf"
    # Check frequency-based classification
    config = data.get("config", {})
    breath_cycle = config.get("breath_cycle", 10)
    if isinstance(breath_cycle, (int, float)):
        if breath_cycle < 5:
            return "fast_resonance"
        elif breath_cycle > 8:
            return "slow_resonance"
    return "resonance"


def get_session_info(filepath: Path) -> Optional[Dict[str, Any]]:
    """Extract session info from a JSON file for listing."""
    data = load_session_data(filepath)
    if not data:
        return None
    
    stat = filepath.stat()
    training_type = detect_training_type(data)
    
    return {
        "filename": filepath.name,
        "id": data.get("id", ""),
        "session_id": data.get("session_id", ""),
        "username": data.get("username", "anonymous"),
        "timestamp": data.get("timestamp", ""),
        "training_type": training_type,
        "training_type_label": get_training_type_label(training_type),
        "size_bytes": stat.st_size,
        "mtime_utc": datetime.fromtimestamp(stat.st_mtime).isoformat(),
        "bpm_avg": data.get("bpm_avg"),
        "bpm_min": data.get("bpm_min"),
        "bpm_max": data.get("bpm_max"),
        "hrv_rmssd": data.get("hrv_rmssd"),
        "lf_power": data.get("lf_power"),
        "rr_count": data.get("rr_count", 0),
    }


def get_training_type_label(training_type: str) -> str:
    """Get human-readable label for training type."""
    labels = {
        "buteyko": "布捷伊科屏气",
        "prbf": "个体共振频率 (PRBF)",
        "fast_resonance": "快速共振呼吸",
        "slow_resonance": "慢速共振呼吸",
        "resonance": "共振呼吸",
    }
    return labels.get(training_type, training_type)


@admin_api_bp.route("/api/admin/sessions")
@admin_required
def list_sessions():
    """
    List all sessions with optional filtering and sorting.
    
    Query params:
        - username: Filter by username (partial match)
        - training_type: Filter by training type
        - sort_by: Sort field (date, username, type, bpm_avg, hrv_rmssd)
        - sort_order: asc or desc
        - limit: Max results (default 100)
    """
    username_filter = request.args.get("username", "").strip().lower()
    training_type_filter = request.args.get("training_type", "").strip()
    sort_by = request.args.get("sort_by", "date")
    sort_order = request.args.get("sort_order", "desc")
    limit = request.args.get("limit", 100, type=int)
    
    # Load all sessions
    files = list(state.DATA_DIR.glob("session_*.json"))
    sessions = []
    for f in files:
        info = get_session_info(f)
        if info:
            sessions.append(info)
    
    # Apply filters
    if username_filter:
        sessions = [s for s in sessions if username_filter in s["username"].lower()]
    
    if training_type_filter:
        sessions = [s for s in sessions if s["training_type"] == training_type_filter]
    
    # Sort
    sort_key_map = {
        "date": "timestamp",
        "username": "username",
        "type": "training_type",
        "bpm_avg": "bpm_avg",
        "hrv_rmssd": "hrv_rmssd",
    }
    sort_key = sort_key_map.get(sort_by, "timestamp")
    reverse = sort_order == "desc"
    
    # Handle None values in sort
    def sort_func(s):
        val = s.get(sort_key)
        if val is None:
            return (1, "") if reverse else (0, "")
        return (0, val)
    
    sessions.sort(key=sort_func, reverse=reverse)
    
    # Limit results
    total = len(sessions)
    sessions = sessions[:limit]
    
    return jsonify({
        "ok": True,
        "sessions": sessions,
        "total": total,
        "filters": {
            "username": username_filter,
            "training_type": training_type_filter,
        },
        "sort": {
            "by": sort_by,
            "order": sort_order,
        },
        "training_types": [
            {"value": "buteyko", "label": "布捷伊科屏气"},
            {"value": "prbf", "label": "个体共振频率 (PRBF)"},
            {"value": "fast_resonance", "label": "快速共振呼吸"},
            {"value": "slow_resonance", "label": "慢速共振呼吸"},
            {"value": "resonance", "label": "共振呼吸"},
        ],
    })


@admin_api_bp.route("/api/admin/sessions/download/<path:filename>")
@admin_required
def download_session(filename: str):
    """Download a single session JSON file."""
    # Validate filename
    if not filename.endswith(".json") or "/" in filename or "\\" in filename or ".." in filename:
        abort(400, description="Invalid filename")
    
    filepath = state.DATA_DIR / filename
    if not filepath.exists() or not filepath.is_file():
        abort(404, description="File not found")
    
    return send_file(
        filepath,
        mimetype="application/json",
        as_attachment=True,
        download_name=filename
    )


@admin_api_bp.route("/api/admin/sessions/download-zip", methods=["POST"])
@admin_required
def download_zip():
    """
    Download multiple sessions as a ZIP file.
    
    Request body:
        - filenames: List of filenames to include
        - Or use filters (same as list_sessions) to auto-select
    """
    data = request.get_json(silent=True) or {}
    filenames = data.get("filenames", [])
    
    # If no filenames provided, use filters
    if not filenames:
        username_filter = data.get("username", "").strip().lower()
        training_type_filter = data.get("training_type", "").strip()
        
        files = list(state.DATA_DIR.glob("session_*.json"))
        for f in files:
            info = get_session_info(f)
            if not info:
                continue
            if username_filter and username_filter not in info["username"].lower():
                continue
            if training_type_filter and info["training_type"] != training_type_filter:
                continue
            filenames.append(info["filename"])
    
    if not filenames:
        return jsonify({"ok": False, "error": "No files selected"}), 400
    
    # Create ZIP in memory
    memory_file = io.BytesIO()
    with zipfile.ZipFile(memory_file, "w", zipfile.ZIP_DEFLATED) as zf:
        for filename in filenames:
            # Validate filename
            if not filename.endswith(".json") or "/" in filename or "\\" in filename or ".." in filename:
                continue
            filepath = state.DATA_DIR / filename
            if filepath.exists() and filepath.is_file():
                zf.write(filepath, filename)
    
    memory_file.seek(0)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    
    return send_file(
        memory_file,
        mimetype="application/zip",
        as_attachment=True,
        download_name=f"heartbeat_sessions_{timestamp}.zip"
    )


@admin_api_bp.route("/api/admin/sessions/stats")
@admin_required
def get_stats():
    """Get overall statistics about stored sessions."""
    files = list(state.DATA_DIR.glob("session_*.json"))
    
    total_files = len(files)
    total_size = sum(f.stat().st_size for f in files)
    usernames = set()
    training_types = {}
    
    for f in files:
        info = get_session_info(f)
        if info:
            usernames.add(info["username"])
            t = info["training_type"]
            training_types[t] = training_types.get(t, 0) + 1
    
    return jsonify({
        "ok": True,
        "stats": {
            "total_sessions": total_files,
            "total_size_bytes": total_size,
            "unique_users": len(usernames),
            "training_types": training_types,
        },
    })


@admin_api_bp.route("/api/admin/sessions/view/<path:filename>")
@admin_required
def view_session(filename: str):
    """View a session JSON file (for preview)."""
    # Validate filename
    if not filename.endswith(".json") or "/" in filename or "\\" in filename or ".." in filename:
        abort(400, description="Invalid filename")
    
    filepath = state.DATA_DIR / filename
    if not filepath.exists() or not filepath.is_file():
        abort(404, description="File not found")
    
    data = load_session_data(filepath)
    if not data:
        abort(500, description="Failed to load session data")
    
    # Add metadata
    stat = filepath.stat()
    data["_admin_meta"] = {
        "filename": filename,
        "size_bytes": stat.st_size,
        "modified": datetime.fromtimestamp(stat.st_mtime).isoformat(),
        "training_type": detect_training_type(data),
    }
    
    # Recalculate HRV metrics from RR intervals for accuracy
    rr_intervals = data.get("rr_intervals_ms", []) or data.get("rr_intervals", [])
    if rr_intervals and len(rr_intervals) >= 30:
        # Recalculate LF Power from RR intervals
        lf_power = compute_lf_power(rr_intervals)
        if lf_power is not None:
            data["_admin_meta"]["lf_power_calculated"] = round(lf_power, 4)
            # Also update the main data for consistency
            data["lf_power"] = round(lf_power, 4)
        
        # Recalculate RMSSD
        rmssd = compute_rmssd(rr_intervals)
        if rmssd is not None:
            data["_admin_meta"]["hrv_rmssd_calculated"] = round(rmssd, 2)
            data["hrv_rmssd"] = round(rmssd, 2)
    
    return jsonify(data)


# ── Export helpers ─────────────────────────────────────────────────────────

def _export_session_to_txt(data: Dict[str, Any]) -> str:
    """Convert session data to tab-separated TXT with header metadata."""
    lines = []
    lines.append("# HeartBeat Session Export")
    lines.append(f"# Session ID: {data.get('id', '')}")
    lines.append(f"# Username: {data.get('username', '')}")
    lines.append(f"# Date: {data.get('timestamp', '')}")
    lines.append(f"# Training Type: {data.get('training_type', '')}")
    lines.append(f"# RR Count: {data.get('rr_count', 0)}")
    bpm_min = data.get('bpm_min', '')
    bpm_avg = data.get('bpm_avg', '')
    bpm_max = data.get('bpm_max', '')
    lines.append(f"# BPM (min/avg/max): {bpm_min} / {bpm_avg} / {bpm_max}")
    lines.append(f"# HRV RMSSD: {data.get('hrv_rmssd', '')} ms")
    lines.append(f"# LF Power: {data.get('lf_power', '')}")
    lines.append("#")
    lines.append("# Columns: timestamp\trr_interval_ms\theart_rate_bpm")
    lines.append("timestamp\trr_interval_ms\theart_rate_bpm")

    rr_intervals = data.get("rr_intervals_ms", []) or data.get("rr_intervals", [])
    rr_timestamps = data.get("rr_timestamps", [])

    for i, rr in enumerate(rr_intervals):
        ts = rr_timestamps[i] if i < len(rr_timestamps) else ""
        if ts is not None and isinstance(ts, (int, float)):
            ts = datetime.fromtimestamp(ts).isoformat()
        hr = round(60000.0 / rr, 1) if rr > 0 else ""
        lines.append(f"{ts}\t{rr}\t{hr}")

    return "\n".join(lines)


def _export_session_to_xlsx(data: Dict[str, Any]) -> io.BytesIO:
    """Convert session data to an Excel workbook with Info + Data sheets."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()

    # Sheet 1: Session Info
    ws_info = wb.active
    ws_info.title = "Session Info"
    ws_info.column_dimensions['A'].width = 20
    ws_info.column_dimensions['B'].width = 45

    label_font = Font(bold=True)
    ws_info.cell(row=1, column=1, value="HeartBeat Session Export").font = Font(bold=True, size=13)

    info_fields = [
        ("Session ID", data.get("id", "")),
        ("Username", data.get("username", "")),
        ("Date", data.get("timestamp", "")),
        ("Training Type", data.get("training_type", "")),
        ("RR Count", data.get("rr_count", 0)),
        ("BPM (min/avg/max)",
         f"{data.get('bpm_min', '')} / {data.get('bpm_avg', '')} / {data.get('bpm_max', '')}"),
        ("HRV RMSSD (ms)", data.get("hrv_rmssd", "")),
        ("LF Power", data.get("lf_power", "")),
    ]
    for i, (label, value) in enumerate(info_fields, start=3):
        ws_info.cell(row=i, column=1, value=label).font = label_font
        ws_info.cell(row=i, column=2, value=str(value) if value is not None else "")

    # Sheet 2: Data
    ws_data = wb.create_sheet("Data")
    hdr_font = Font(bold=True, color="FFFFFF")
    hdr_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    for col, (name, width) in enumerate(
        [("timestamp", 22), ("rr_interval_ms", 16), ("heart_rate_bpm", 16)], start=1
    ):
        cell = ws_data.cell(row=1, column=col, value=name)
        cell.font = hdr_font
        cell.fill = hdr_fill
        ws_data.column_dimensions[get_column_letter(col)].width = width

    rr_intervals = data.get("rr_intervals_ms", []) or data.get("rr_intervals", [])
    rr_timestamps = data.get("rr_timestamps", [])

    for i, rr in enumerate(rr_intervals):
        row = i + 2
        ts = rr_timestamps[i] if i < len(rr_timestamps) else ""
        if ts is not None and isinstance(ts, (int, float)):
            ts = datetime.fromtimestamp(ts).isoformat()
        ws_data.cell(row=row, column=1, value=ts)
        ws_data.cell(row=row, column=2, value=rr)
        if rr > 0:
            ws_data.cell(row=row, column=3, value=round(60000.0 / rr, 1))

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def _build_export(data: Dict[str, Any], fmt: str, base_name: str):
    """Return (BytesIO, mimetype, download_name) for the chosen format."""
    if fmt == "txt":
        content = _export_session_to_txt(data).encode("utf-8")
        return io.BytesIO(content), "text/plain; charset=utf-8", f"{base_name}.txt"
    elif fmt == "xlsx":
        return _export_session_to_xlsx(data), \
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", \
            f"{base_name}.xlsx"
    else:
        content = json.dumps(data, ensure_ascii=False, indent=2).encode("utf-8")
        return io.BytesIO(content), "application/json; charset=utf-8", f"{base_name}.json"


# ── Export endpoints ───────────────────────────────────────────────────────

@admin_api_bp.route("/api/admin/sessions/export/<path:filename>")
@admin_required
def export_session(filename: str):
    """Export a single session in txt, xlsx, or json format."""
    fmt = request.args.get("format", "json").lower()
    if fmt not in ("txt", "xlsx", "json"):
        abort(400, description="Invalid format. Use: txt, xlsx, json")

    if not filename.endswith(".json") or "/" in filename or "\\" in filename or ".." in filename:
        abort(400, description="Invalid filename")

    filepath = state.DATA_DIR / filename
    if not filepath.exists() or not filepath.is_file():
        abort(404, description="File not found")

    data = load_session_data(filepath)
    if not data:
        abort(500, description="Failed to load session data")

    base_name = filename.rsplit(".", 1)[0]
    content, mimetype, download_name = _build_export(data, fmt, base_name)

    return send_file(
        content,
        mimetype=mimetype,
        as_attachment=True,
        download_name=download_name,
    )


@admin_api_bp.route("/api/admin/sessions/export-zip", methods=["POST"])
@admin_required
def export_zip():
    """
    Download multiple sessions as a ZIP file in the selected format.

    Request body:
        - filenames: List of filenames to include
        - format: "txt", "xlsx", or "json" (default "json")
        - Or use filters to auto-select
    """
    body = request.get_json(silent=True) or {}
    filenames = body.get("filenames", [])
    fmt = body.get("format", "json").lower()

    if fmt not in ("txt", "xlsx", "json"):
        return jsonify({"ok": False, "error": "Invalid format"}), 400

    if not filenames:
        username_filter = body.get("username", "").strip().lower()
        training_type_filter = body.get("training_type", "").strip()

        files = list(state.DATA_DIR.glob("session_*.json"))
        for f in files:
            info = get_session_info(f)
            if not info:
                continue
            if username_filter and username_filter not in info["username"].lower():
                continue
            if training_type_filter and info["training_type"] != training_type_filter:
                continue
            filenames.append(info["filename"])

    if not filenames:
        return jsonify({"ok": False, "error": "No files selected"}), 400

    memory_file = io.BytesIO()
    with zipfile.ZipFile(memory_file, "w", zipfile.ZIP_DEFLATED) as zf:
        for filename in filenames:
            if not filename.endswith(".json") or "/" in filename or "\\" in filename or ".." in filename:
                continue
            filepath = state.DATA_DIR / filename
            if not filepath.exists() or not filepath.is_file():
                continue
            data = load_session_data(filepath)
            if not data:
                continue
            base_name = filename.rsplit(".", 1)[0]
            content, _, export_name = _build_export(data, fmt, base_name)
            zf.writestr(export_name, content.read())

    memory_file.seek(0)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    return send_file(
        memory_file,
        mimetype="application/zip",
        as_attachment=True,
        download_name=f"heartbeat_export_{timestamp}.zip",
    )
